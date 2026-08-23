import openpyxl
from django.db import transaction

from api.views import get_unit_pk, get_user_pk
from app.models import Estimate, Task

parent_id = {}


@transaction.atomic
def upload_excel_estimate(excel_file, form):
    Task.objects.all().delete()

    wb = openpyxl.load_workbook(excel_file)
    estimate_new_id = None
    for i, worksheet in enumerate(wb.worksheets):

        if i == 0:
            # 1枚目のシート（インデックス 0）の処理
            estimate_new_id = first_sheet(worksheet, form)
        else:
            # 2枚目以降のシートの処理
            after_sheet(worksheet, form, estimate_new_id)
            print(i)
    # 関数の最後にIDを返してあげる（これで警告が消えます）
    # return estimate_new_id


def first_sheet(worksheet, form):
    # --- Estimate情報 ---
    estimate_obj = {}
    task_obj = {}
    # 事業者PK
    client = form.cleaned_data['client_pk']
    estimate_obj["client"] = client

    # Formで選択した年度
    fiscalyear = form.cleaned_data['fiscalyear']
    estimate_obj["fiscalyear"] = fiscalyear.pk

    # 見積年月日
    estimate_date = worksheet['AA7'].value
    formatted_date = estimate_date.strftime("%Y年%m月%d日")
    estimate_obj["estimate_date"] = formatted_date

    # 見積書印刷年月日
    estimate_print_date = worksheet['AA7'].value
    formatted_date = estimate_print_date.strftime("%Y年%m月%d日")
    estimate_obj["estimate_print_date"] = formatted_date

    # Formから新規に登録した見積書番号
    estimate_no = form.cleaned_data['estimate_no']
    estimate_obj["estimate_no"] = estimate_no

    # 発注者名１
    orderer_name1 = worksheet['AN9'].value
    estimate_obj["orderer_name1"] = orderer_name1

    # 発注者名２
    orderer_name2 = worksheet['AO9'].value
    estimate_obj["orderer_name2"] = orderer_name2

    # 税込請負金額
    estimate_amount = worksheet['F9'].value
    estimate_obj["estimate_amount"] = estimate_amount

    # 税区分
    consumption_cls = worksheet['AN5'].value
    estimate_obj["consumption_cls"] = consumption_cls

    # 工事件名
    estimate_name = worksheet['F12'].value
    estimate_obj["estimate_name"] = estimate_name

    # 工事現場住所１
    contract_address1 = worksheet['AN11'].value
    estimate_obj["contract_address1"] = contract_address1

    # 工事現場住所２
    contract_address2 = worksheet['AO11'].value
    estimate_obj["contract_address2"] = contract_address2

    # 工事現場住所３
    contract_address3 = worksheet['AP11'].value
    estimate_obj["contract_address3"] = contract_address3

    # 見積有効期限
    estimate_limit_date = worksheet['F16'].value
    if estimate_limit_date == '0000/00/00':
        estimate_limit_date = None
    estimate_obj["estimate_limit_date"] = estimate_limit_date

    # 支払条件
    payment_term = worksheet['F18'].value
    estimate_obj["payment_term"] = payment_term

    # 工期又は納期
    estimate_end_date = worksheet['F20'].value
    if estimate_end_date == '0000/00/00':
        estimate_end_date = None
    estimate_obj["estimate_end_date"] = estimate_end_date

    # 受渡場所
    delivery_location = worksheet['F22'].value
    estimate_obj["delivery_location"] = delivery_location

    # markup_rate ?  取れない 100にするか検討

    # 工事担当者
    estimate_person_name = worksheet['V16'].value
    clientPk = form.cleaned_data['client_pk']
    estimate_personPk = get_user_pk(clientPk, estimate_person_name)
    estimate_obj["estimate_person"] = estimate_personPk

    # Formから選択した得意先
    customer = form.cleaned_data['customer']
    estimate_obj["customer"] = customer.pk

    # Estimateに書き込み親のIDを取得
    estimate_new_id = write_estimate(estimate_obj)

    # --- Taskの取得 ---
    row_counter = worksheet.max_row
    estimate_tax_amount = None
    for i in range(row_counter - 28):

        task_obj["estimate_id"] = estimate_new_id

        material_dimensions = worksheet.cell(row=i + 27, column=10).value
        if material_dimensions is None:
            material_dimensions = ""
        task_obj["material_dimensions"] = material_dimensions

        quantity = worksheet.cell(row=i + 27, column=17).value
        if quantity is None:
            quantity = None
        task_obj["quantity"] = quantity

        unit_name = worksheet.cell(row=i + 27, column=20).value
        clientPk = form.cleaned_data['client_pk']
        unitPk = get_unit_pk(clientPk, unit_name)
        task_obj["unit"] = unitPk

        price = worksheet.cell(row=i + 27, column=22).value
        if price is None:
            price = None
        task_obj["price"] = price

        amount = worksheet.cell(row=i + 27, column=26).value
        if amount is None:
            amount = None
        task_obj["amount"] = amount

        note = worksheet.cell(row=i + 27, column=30).value
        if note is None:
            note = ""
        task_obj["note"] = note

        aggregation = worksheet.cell(row=i + 27, column=36).value
        if aggregation is None:
            aggregation = ""
        task_obj["aggregation"] = aggregation

        task_name = worksheet.cell(row=i + 27, column=2).value
        if task_name == '消費税':
            estimate_tax_amount = amount
        if task_name is None:
            task_name = ""
        task_obj["task_name"] = task_name

        task_obj["parent"] = None

        parent = write_task(task_obj)
        parent_id[task_name] = parent

        print(parent_id)

        estimate = Estimate.objects.get(pk=estimate_new_id)
        estimate.estimate_tax_amount = estimate_tax_amount
        estimate.save()

        # estimate_tax_amount　明細を先に読んで消費税の金額をセットする
        # estimate_obj["estimate_tax_amount"] = estimate_tax_amount

        # print(task_name, material_dimensions, quantity, unitPk, price, amount, note, aggregation, parent)

    return estimate_new_id


def write_estimate(estimate_obj):
    estimate = Estimate(
        client_id=estimate_obj["client"],
        fiscalyear_id=estimate_obj["fiscalyear"],
        estimate_date=estimate_obj["estimate_date"],
        estimate_print_date=estimate_obj["estimate_print_date"],
        estimate_no=estimate_obj["estimate_no"],
        orderer_name1=estimate_obj["orderer_name1"],
        orderer_name2=estimate_obj["orderer_name2"],
        estimate_amount=estimate_obj["estimate_amount"],
        # estimate_tax_amount=estimate_obj["estimate_tax_amount"],
        consumption_cls=estimate_obj["consumption_cls"],
        estimate_name=estimate_obj["estimate_name"],
        contract_address1=estimate_obj["contract_address1"],
        contract_address2=estimate_obj["contract_address2"],
        estimate_limit_date=estimate_obj["estimate_limit_date"],
        payment_term=estimate_obj["payment_term"],
        estimate_end_date=estimate_obj["estimate_end_date"],
        delivery_location=estimate_obj["delivery_location"],
        estimate_person_id=estimate_obj["estimate_person"],
        customer_id=estimate_obj["customer"],

    )
    estimate.save()

    return estimate.id


def write_task(task_obj):
    task = Task(
        estimate_id=task_obj["estimate_id"],
        task_name=task_obj["task_name"],
        material_dimensions=task_obj["material_dimensions"],
        # budget_quantity=task_obj["budget_quantity"],
        # budget_unit=task_obj["budget_unit"],
        # budget_name=task_obj["budget_name"],
        # budget_amount=task_obj["budget_amount"],
        quantity=task_obj["quantity"],
        unit_id=task_obj["unit"],
        price=task_obj["price"],
        amount=task_obj["amount"],
        # markup_rate=task_obj["markup_rate"],
        # aggregation=task_obj["aggregation"],
        note=task_obj["note"],
        parent_id=task_obj["parent"],
        # sort_order=task_obj["sort_order"],
    )
    task.save()

    return task.id


def after_sheet(worksheet, form, estimate_new_id):
    row_counter = worksheet.max_row
    task_obj = {}
    for i in range(row_counter - 5):

        task_obj["estimate_id"] = estimate_new_id
        task_name = worksheet.cell(row=i + 4, column=2).value
        if task_name is None:
            task_name = ""
        task_obj["task_name"] = task_name

        material_dimensions = worksheet.cell(row=i + 4, column=3).value
        if material_dimensions is None:
            material_dimensions = ""
        task_obj["material_dimensions"] = material_dimensions

        quantity = worksheet.cell(row=i + 4, column=4).value
        # if quantity is None:
        #     quantity = ""
        task_obj["quantity"] = quantity

        unit_name = worksheet.cell(row=i + 4, column=5).value
        clientPk = form.cleaned_data['client_pk']
        unitPk = get_unit_pk(clientPk, unit_name)
        task_obj["unit"] = unitPk

        price = worksheet.cell(row=i + 4, column=6).value
        # if price is None:
        #     price = ""
        task_obj["price"] = price

        amount = worksheet.cell(row=i + 4, column=7).value
        # if amount is None:
        #     amount = ""
        task_obj["amount"] = amount

        note = worksheet.cell(row=i + 4, column=8).value
        if note is None:
            note = ""
        task_obj["note"] = note

        aggregation = worksheet.cell(row=i + 4, column=10).value
        if aggregation is None:
            aggregation = ""
        task_obj["aggregation"] = aggregation

        parent_task_name = worksheet.cell(row=2, column=2).value
        parent_task_id = parent_id.get(parent_task_name)

        task_obj["parent"] = parent_task_id

        write_task(task_obj)
        # print(task_name, material_dimensions, quantity, unitPk, price, amount, note, aggregation)
